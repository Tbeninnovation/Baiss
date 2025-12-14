import pandas as pd
import os
from typing import List, Dict, Any
import sys
import io

import openpyxl
from collections import deque


import baisstools
baisstools.insert_syspath(__file__, matcher=[r"^baiss_.*$"])

from baiss_sdk.parsers.base_parser import BaseParser
from baiss_sdk.files.file_reader import FileReader
class ExcelParser(BaseParser):
    """
    An Excel parser that uses openpyxl to detect data blocks within each sheet
    and pandas to process them as structured or unstructured content.
    """
    def __init__(self):
        """Initializes the parser."""
        super().__init__()
        print("Block-based Excel Parser initialized.")

    def detect_data_blocks(self, sheet) -> List[pd.DataFrame]:
        """
        Scans a worksheet to identify contiguous blocks of non-empty cells using BFS.

        Args:
            sheet: An openpyxl worksheet object.

        Returns:
            A list of pandas DataFrames, each representing a data block.
        """
        if not sheet.max_row or not sheet.max_column:
            return []

        visited = set()
        blocks = []

        for row_idx in range(1, sheet.max_row + 1):
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if (row_idx, col_idx) in visited or cell.value is None or str(cell.value).strip() == "":
                    continue

                # Start BFS for a new block
                q = deque([(row_idx, col_idx)])
                visited.add((row_idx, col_idx))
                block_cells = []
                min_r, max_r = row_idx, row_idx
                min_c, max_c = col_idx, col_idx

                while q:
                    r, c = q.popleft()
                    block_cells.append((r, c))
                    min_r, max_r = min(min_r, r), max(max_r, r)
                    min_c, max_c = min(min_c, c), max(max_c, c)

                    # Explore neighbors (up, down, left, right)
                    for dr, dc in [(0, 1), (0, -1), (1, 0), (-1, 0)]:
                        nr, nc = r + dr, c + dc
                        if 1 <= nr <= sheet.max_row and 1 <= nc <= sheet.max_column:
                            neighbor_cell = sheet.cell(row=nr, column=nc)
                            if (nr, nc) not in visited and neighbor_cell.value is not None and str(neighbor_cell.value).strip() != "":
                                visited.add((nr, nc))
                                q.append((nr, nc))

                # Create DataFrame from the detected block
                if not block_cells:
                    continue

                data = [[sheet.cell(row=r, column=c).value for c in range(min_c, max_c + 1)] for r in range(min_r, max_r + 1)]
                df = pd.DataFrame(data)

                # Use the first row as header
                if not df.empty:
                    df.columns = df.iloc[0]
                    df = df[1:]
                    df.reset_index(drop=True, inplace=True)

                blocks.append(df)

        return blocks

    def parse(self, excel_path: str, max_tokens_per_chunk: int = 500) -> List[Dict[str, Any]]:
        """
        Reads an Excel file, detects data blocks on each sheet, processes them, and creates chunks.
        """
        excel_path = FileReader.update_file_path(excel_path)
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel file not found at: {excel_path}")

        try:
            # Handle .xls files by converting them to an in-memory .xlsx representation
            if excel_path.lower().endswith('.xls'):
                print("DEBUG: .xls file detected. Converting to in-memory .xlsx for processing.")
                xls_file = pd.ExcelFile(excel_path, engine='xlrd')
                excel_sheets = {sheet_name: xls_file.parse(sheet_name, header=None) for sheet_name in xls_file.sheet_names}

                in_memory_xlsx = io.BytesIO()
                with pd.ExcelWriter(in_memory_xlsx, engine='openpyxl') as writer:
                    for sheet_name, df in excel_sheets.items():
                        df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

                in_memory_xlsx.seek(0)
                workbook = openpyxl.load_workbook(in_memory_xlsx, data_only=True)
            else:
                # Load .xlsx files directly
                workbook = openpyxl.load_workbook(excel_path, data_only=True)

            all_results = []

            for sheet_num, sheet_name in enumerate(workbook.sheetnames):
                sheet = workbook[sheet_name]
                print(f"\nProcessing sheet '{sheet_name}'...")

                data_blocks = self.detect_data_blocks(sheet)
                if not data_blocks:
                    # print(f"No data blocks found in sheet '{sheet_name}'.")
                    continue

                # print(f"Found {len(data_blocks)} data block(s) in sheet '{sheet_name}'.")

                for block_index, df in enumerate(data_blocks):
                    df.dropna(how='all', inplace=True)
                    df.dropna(how='all', axis=1, inplace=True)
                    if df.empty:
                        continue

                    print(f"  - Processing block {block_index} with shape {df.shape}")
                    is_block_structured = self.is_structured(df)
                    chunks = []

                    if is_block_structured:
                        print(f"    Block {block_index} is structured. Processing as a table.")
                        # Re-add header for markdown conversion
                        df_with_header = pd.concat([pd.DataFrame([df.columns], columns=df.columns), df], ignore_index=True)
                        markdown_content = df_with_header.to_markdown(index=False, tablefmt="pipe")

                        if markdown_content and markdown_content.strip():
                            markdown_lines = markdown_content.split('\n')
                            header_line = markdown_lines[0]
                            separator_line = markdown_lines[1]
                            header_with_separator = f"{header_line}\n{separator_line}"

                            data_rows = [row.strip().split('|')[1:-1] for row in markdown_lines[2:]]
                            data_rows_cleaned = [[cell.strip() for cell in row] for row in data_rows]

                            row_token_data = self._calculate_row_tokens(data_rows_cleaned, header_with_separator)
                            chunks = self._create_chunks_by_tokens(header_with_separator, row_token_data, max_tokens_per_chunk, sheet_name=sheet_name)
                    else:
                        print(f"    Block {block_index} is unstructured. Processing as text.")
                        raw_text = df.to_string(index=False, header=False)
                        chunks = self._create_chunks_from_text(raw_text, max_tokens_per_chunk, source=excel_path, sheet_name=sheet_name)

                    for i, chunk in enumerate(chunks):
                        all_results.append({
                            "content": chunk["content"],
                            "metadata": {
                                "source": excel_path,
                                "sheet_name": sheet_name,
                                "sheet_number": sheet_num,
                                "block_index": block_index,
                                "chunk_index": i,
                                "tokens": chunk["tokens"],
                                "is_structured": is_block_structured
                            }
                        })
            return all_results
        except Exception as e:
            raise ValueError(f"Failed to parse Excel file with openpyxl/pandas: {e}")


if __name__ == "__main__":
    pass
